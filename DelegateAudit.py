#!/usr/bin/env python3

import sqlite3

import openpyxl
from openpyxl.cell import coordinate_from_string, column_index_from_string
from openpyxl import load_workbook

def load_database(db_name):
    conn = sqlite3.connect(db_name)
    cur = conn.cursor()
    print(conn,cur)
    return([conn, cur])
    
def setup_database():
    [conn, cur] = load_database('/Users/sroberts/Dropbox/TMT/Python/Sandbox/DelegatedAuthorities.sqlite')
    cur.execute('DROP TABLE IF EXISTS configuration_management ')
    cur.execute('CREATE TABLE configuration_management( \
        Number integer, \
        Role string, \
        Position string, \
        Authority string, \
        Alternate string, \
        PRIMARY KEY (Number))')
        
    cur.execute('DROP TABLE IF EXISTS board ')
    cur.execute('CREATE TABLE board( \
        Role string, \
        Function string, \
        Name string, \
        Email string, \
        Alternate string, \
        Assnt_name string, \
        Assnt_email string)')        

    return([conn, cur])
    
    
def ss_board_to_db(conn, cur, ws, start_rc, end_rc):
    data = get_ss_data(ws,start_rc,end_rc)
    for d in data:
        print('Email is :', d[3])
        if d[3]:
            print('Entry is :', d)
            cur.execute('INSERT INTO board (Role, Function, Name, Email, Alternate, Assnt_name, Assnt_email) VALUES ( ?,?,?,?,?,?,? )', ( d[0], d[1], d[2], d[3], d[4], d[5], d[6] )) 
    conn.commit() 

def get_row_col_from_string(string):
    xy = coordinate_from_string(string) # returns ('A',4)
    col = column_index_from_string(xy[0]) # returns 1
    row = xy[1]
    return([row, col])

def get_ss_data(ws, start_rc, end_rc):
    data = []
    [srow, scol] = get_row_col_from_string(start_rc)
    [erow, ecol] = get_row_col_from_string(end_rc)
    for row in range(srow,erow+1):
        rowvalues = []
        for col in range(scol,ecol+1):
            entry = ws.cell(row = row, column = col).value
            rowvalues.append(entry)
        data.append(rowvalues)
    return(data)
    
def ss_ccb_to_db(conn, cur, ws, start_rc, end_rc):
    data = get_ss_data(ws,start_rc,end_rc)
    for d in data:
        print('Entry is :', d)
        cur.execute('INSERT INTO configuration_management (Number, Role, Position, Authority, Alternate) VALUES ( ?,?,?,?,? )', ( d[0], d[1], d[2], d[3], d[4] )) 
    conn.commit() 

if __name__ == '__main__':
    [conn, cur] = setup_database()

    xlfile = '/Users/sroberts/Dropbox/TMT/Python/Sandbox/20160913_Listing of Delegated Authorities_FTv03_ACMv04_TMT BUS MGT 15 176 REL10.xlsx'
    # xlfile = 'test.xlsx'
    wb = load_workbook(xlfile)
    print('Opened existing file :', xlfile)
    ws = wb.worksheets[3]

    ss_ccb_to_db(conn, cur, ws, 'A6', 'E16')
    
    ws = wb.worksheets[6]
    ss_board_to_db(conn, cur, ws, 'C5', 'I56')
    
